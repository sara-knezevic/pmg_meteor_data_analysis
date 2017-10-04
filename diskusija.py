import csv, os, shutil, sys
from openpyxl import *
from openpyxl.styles import *
import time, datetime

def createFolders(source, destination):
	for root, dirs, files in os.walk(source):
		for file in files:
			folder = file[1:9]

			if not os.path.exists(os.path.join(destination, folder)):
				os.makedirs(os.path.join(destination, folder))

def sortFiles(source, destination):
	for root, dirs, files in os.walk(source):
		for file in files:
			name = file[1:9]
			source_file = os.path.join(root, file)
			destination_file = os.path.join(destination, name)

			shutil.copy(source_file, destination_file)

def createExcelFile(directory):
	for root, dirs, files in os.walk(directory):
		wb = Workbook()
		ws = wb.active

		name = root[-8:]
		excel = name + ".xlsx"

		wb.save(os.path.join(root, excel))

def storeInExcel(directory):
	for root, dirs, files in os.walk(directory):
		for file in files:

			wb = load_workbook(os.path.join(root, root[-8:] + ".xlsx"))
			global ws
			ws = wb.active

			if (file[-4:] != "xlsx"):
				with open(os.path.join(root, file), "rb") as csvFile:
					read = csv.reader(csvFile)

					for row in read:
						if None in row:
							continue
						elif "Ver" in row:
							continue
						else:
							ws.append(row)
							
			wb.save(os.path.join(root, root[-8:] + ".xlsx"))

def convertTime(directory):
	for root, dirs, files in os.walk(directory):
		wb = load_workbook(os.path.join(root, root[-8:] + ".xlsx"))
		global ws
		ws = wb.active

		r = 2
		c = 3

		for row in ws.iter_rows("C{}:C{}".format(ws.min_row, ws.max_row)):
			year = str(cell(r, 9))
			month = str(cell(r, 10))
			day = str(cell(r, 11))
			hour = str(cell(r, 12))
			minute = str(cell(r, 13))
			second = str(cell(r, 14))

			if "." in second:
				second = second[:-4]

			date = str(year + "/" + month + "/" + day + " " + hour + ":" + minute + ":" + second)
			ws.cell(row = r, column = c).value = date

			wb.save(os.path.join(root, root[-8:] + ".xlsx"))
			r += 1

def seconds(date):
	return date.tm_sec + date.tm_min * 60 + date.tm_hour * 3600 + date.tm_mday * 24 * 3600

def cell(r, c):
	return str(ws.cell(row = r, column = c).value)

def compare(destination):
	for root, dirs, files in os.walk(destination):
		name = root[-8:]
		excelfile = name + ".xlsx"

		wb = load_workbook(os.path.join(root, excelfile))
		global ws
		ws = wb.active

		timeFormat = "%Y/%m/%d %H:%M:%S"

		for file in files: # one excel file

			if file[-4:] == "xlsx":

				print "-	-	-	-	-	-	-	-	-	-"
				print file[:-5]
				print "-	-	-	-	-	-	-	-	-	-"

				r = 2
				times = []

				for row in ws.iter_rows("C{}:C{}".format(ws.min_row + 1, ws.max_row)):
					d = time.strptime(cell(r, 3), timeFormat)
					times.append(d)

					r += 1

				counter = len(times)
				index = 0

				for i in xrange(0, counter):
					for j in xrange(counter - 1, i, -1):
						if (abs(seconds(times[i]) - seconds(times[j])) < 2):
							if (cell(i + 2, 7) != cell(j + 2, 7)):
								print cell(i + 2, 7) + " " + cell(i + 2, 3)
								print "Group =" + cell(i + 2, 2) + "	" + "mag = " + cell(i + 2, 4) + "	" + "dur = " + cell(i + 2, 5) + "\n"

								print cell(j + 2, 7) + " " + cell(j + 2, 3)
								print "Group =" + cell(j + 2, 2) + "	" + "mag = " + cell(j + 2, 4) + "	" + "dur = " + cell(j + 2, 5) + "\n"


				times = []
				wb.save(os.path.join(root, excelfile))

source = "C:\Users\saras\Desktop\PMG\CSV"
destination = "C:\Users\saras\Desktop\Norveska"

createFolders(source, destination)
sortFiles(source, destination)
createExcelFile(destination)
storeInExcel(destination)
convertTime(destination)
compare(destination)
