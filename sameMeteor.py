
import os, shutil, sys
import time, datetime
from openpyxl import *

def createExcel(directory, filename):
	wb = Workbook()
	ws = wb.active
	wb.save(os.path.join(directory, filename))

def seconds(date):
	return date.tm_sec + date.tm_min * 60 + date.tm_hour * 3600 + date.tm_mday * 24 * 3600

def storingTime(directory, name, date):

	baza = load_workbook(os.path.join(directory, name))
	worksheet = baza[date] # name of worksheet, access

	global times # everything storage
	times = []

	global timeFormat
	timeFormat = "%Y-%m-%d %H:%M:%S"

	group = worksheet.cell(row = 3, column = 1).value # first group

	groupRow = 3 # row where the group and observers are
	iterRow = 4 # row variable to iterate through entire table

	# storing times into array

	for row in worksheet.iter_rows(min_row = 3):
		if (worksheet.cell(row = iterRow, column = 4).value is None):
			# time doesn't exist
			if (worksheet.cell(row = iterRow, column = 1).value is not None):
				group = worksheet.cell(row = iterRow, column = 1).value # set group
				groupRow = iterRow # sets row for observers

		else:
			# where the magic happens
			# *cracks knuckles*

			oneTime = []
			oneTime.append(group) # set which group is being inserted

			dateAndTime = str(worksheet.cell(row = iterRow, column = 1).value)[:10] + " " + str(worksheet.cell(row = iterRow, column = 4).value)
			dateAndTime = time.strptime(dateAndTime, timeFormat)

			oneTime.append(dateAndTime) # place time into one array of array

			# 40 column is the last observer, so iterate to 40
			observerRow = iterRow - (iterRow - groupRow)

			kolonaZaMeteore = 10

			posmatraci = {}

			while (kolonaZaMeteore <= 40):
				meteorInfo = []

				if (worksheet.cell(row = iterRow, column = kolonaZaMeteore).value is None):
					kolonaZaMeteore += 1
				else:
					meteor = worksheet.cell(row = iterRow, column = kolonaZaMeteore + 1).value
					magnituda = worksheet.cell(row = iterRow, column = kolonaZaMeteore).value
					posmatrac = worksheet.cell(row = observerRow, column = kolonaZaMeteore).value

					meteorInfo.append(meteor)
					meteorInfo.append(magnituda)

					posmatraci[posmatrac] = meteorInfo

					kolonaZaMeteore += 2

			oneTime.append(posmatraci)

			if bool(oneTime[2]): # ako nije random time stamp bez posmatraca
				times.append(oneTime)

		iterRow += 1

def write(date):

	file = open("Kodovi " + date + ".txt", "w")

	for i in range(0, len(times)):

		file.write("\n")

		file.write(time.strftime(timeFormat, times[i][1]) + "\n")

		file.write("\n")

		file.write(times[i][0] + ": \n")

		for key, value in times[i][2].items():
			file.write(str(key) + " -> " + str(value) + "\n")

		for j in range(i + 1, len(times)):
			comparer = seconds(times[i][1])
			comparee = seconds(times[j][1])

			if (abs(comparer - comparee) <= 3 and times[i][0] != times[j][0]):

				file.write("\n")

				file.write(times[j][0] + ": \n")

				for key, value in times[j][2].items():
					file.write(str(key) + " -> " + str(value) + "\n")

				file.write("\n")

	file.close()
