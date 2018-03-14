# ovo kao final treba da bude
# sve ovde da se smesti
# checked and done

import os, shutil, sys
from openpyxl import *
import numpy as np
import time, datetime

def convert_to_seconds(date):
	return date.tm_sec + date.tm_min * 60 + date.tm_hour * 3600 + date.tm_mday * 24 * 3600

def store_observers(sheet): # stores all observers for one night
	global observers
	observers = set()

	R = 1

	for row in sheet.iter_rows(min_row = 1):
		observers.add(sheet.cell(row = R, column = 7).value)
		R += 1

	observers = list(observers)
	return observers

def cell(worksheet, r, c):
	return worksheet.cell(row = r, column = c).value

def timings(posmatraci, vremena): # stores times of observations into a set of non repeating values
	time_format = "%Y/%m/%d %H:%M:%S"
	day, night = date.split("-")

	time_of_observers = {}

	iter_row = 1
	for row in vremena.iter_rows(min_row = 1):
		o = vremena.cell(row = iter_row, column = 1).value

		tx = ""
		ty = ""

		# starts
		if (int(cell(vremena, iter_row, 2)[:2]) > 17):
			tx = "2017/08/" + day + " " + cell(vremena, iter_row, 2)
			# ends
			if (int(cell(vremena, iter_row, 3)[:2]) > 17):
				ty = "2017/08/" + day + " " + cell(vremena, iter_row, 3)
			else:
				ty = "2017/08/" + night + " " + cell(vremena, iter_row, 3)

		else:
			tx = "2017/08/" + night + " " + cell(vremena, iter_row, 2)
			# ends
			if (int(cell(vremena, iter_row, 3)[:2]) > 17):
				ty = "2017/08/" + day + " " + cell(vremena, iter_row, 3)
			else:
				ty = "2017/08/" + night + " " + cell(vremena, iter_row, 3)

		t_start = time.strptime(tx[:-1], time_format)
		t_end = time.strptime(ty[:-1], time_format)

		time_of_observers.setdefault(o, []).append(t_start)
		time_of_observers.setdefault(o, []).append(t_end)

		iter_row += 1

	return time_of_observers

def compare_meteors(sheet):
	meteors = {}
	time_format = "%Y-%m-%d %H:%M:%S"

	meteors.setdefault("Istok", {})
	meteors.setdefault("Sever", {})
	meteors.setdefault("Zapad", {})

	r = 1

	for row in sheet.iter_rows(min_row = 1):

		group = cell(sheet, r, 5)
		observer = cell(sheet, r, 7)
		lmg = cell(sheet, r, 8)
		mag = cell(sheet, r, 9)
		met = cell(sheet, r, 10)

		t = str(cell(sheet, r, 1))[:-8] + str(cell(sheet, r, 3))
		t = time.strptime(t, time_format)

		# group		time	 observer	 magnitude 		meteor		lmg
		meteors.setdefault(group, {}).setdefault(t, {}).setdefault(observer, []).append(mag)
		meteors.setdefault(group, {}).setdefault(t, {}).setdefault(observer, []).append(met)
		meteors.setdefault(group, {}).setdefault(t, {}).setdefault(observer, []).append(lmg)
		meteors.setdefault(group, {}).setdefault(t, {}).setdefault(observer, []).append(group)

		r += 1

	# ^^^ ovo valja ^^^

	global compared_times
	compared_times = {}

	# to be compared :
	# I -> Z -> S

	# u kurac visess
	for keyIstok, valueIstok in meteors["Istok"].items():
		compared_times.setdefault(keyIstok, []).append(valueIstok)

		for keyZapad, valueZapad in meteors["Zapad"].items():
			if (abs(convert_to_seconds(keyIstok) - convert_to_seconds(keyZapad)) > 2):
				compared_times.setdefault(keyZapad, []).append(valueZapad)

			for keySever, valueSever in meteors["Sever"].items():
				if ((abs(convert_to_seconds(keyIstok) - convert_to_seconds(keySever)) > 2) or (abs(convert_to_seconds(keySever) - convert_to_seconds(keyZapad)) > 2)):
					compared_times.setdefault(keySever, []).append(valueSever)

	return 0

def store_in_excel(p, t):
	wb = Workbook()
	ws = wb.active

	# stores observers
	n = 2
	for i in range(0, len(observers)):
		ws.merge_cells(start_row = 1, start_column = n, end_row = 1, end_column = n + 4)
		ws.cell(row = 1, column = n).value = observers[i]
		n += 5

	# stores meteor time stamps
	time_format = "%Y-%m-%d %H:%M:%S"

	n = 2

	for keyTime, valueTime in compared_times.items():
		ws.cell(row = n, column = 1).value = time.strftime(time_format, keyTime)

		for i in range(0, len(valueTime)):
			for k, v in valueTime[i].items():
				position_column = observers.index(k) * 5 + 2
				ws.cell(row = n, column = position_column).value = v[0] # stores magnitude
				ws.cell(row = n, column = position_column + 1).value = v[1] # stores shower
				ws.cell(row = n, column = position_column + 2).value = v[2] # stores lmg
				ws.cell(row = n, column = position_column + 3).value = v[3] # stores group

		n += 1

	wb.save("storage.xlsx")

	return 0

def check_if_observed(p_ws, t_ws):
	wb = load_workbook("/home/saras/Downloads/kodovi-pmg-master/storage.xlsx")
	ws = wb.active

	time_format = "%Y-%m-%d %H:%M:%S"

	for row in range(2, len(compared_times) + 2): 	 # iterates rows
		t_of_meteor = time.strptime(ws.cell(row = row, column = 1).value, time_format)

		for col in range(6, len(observers) * 5 + 2, 5):		# iterate columns

			posmatrac = ws.cell(row = row - (row - 1), column = col - 4).value
			les_temps = timings(p_ws, t_ws)

			if posmatrac in les_temps:
				for x in range(0, len(les_temps[posmatrac]) - 1, 2):
					start = les_temps[posmatrac][x]
					end = les_temps[posmatrac][x + 1]

					if ((convert_to_seconds(start) < convert_to_seconds(t_of_meteor)) and ((convert_to_seconds(end) > convert_to_seconds(t_of_meteor)))):
						ws.cell(row = row, column = col).value = "P"
					elif (ws.cell(row = row, column = col).value != "P"):
						ws.cell(row = row, column = col).value = "NP"

	wb.save("/home/saras/Downloads/kodovi-pmg-master/Finals/" + date + ".xls")

	return 0

posmatranja = load_workbook("/home/saras/Downloads/kodovi-pmg-master/posmatranja.xlsx")
vremena = load_workbook("/home/saras/Downloads/kodovi-pmg-master/timings.xlsx")
dates = ["8-9", "9-10", "10-11", "11-12", "14-15", "15-16", "16-17", "17-18"]

for d in range(0, len(dates)):
	date = dates[d]
	p_ws = posmatranja[date]
	t_ws = vremena[date]

	observers = store_observers(p_ws)

	compare_meteors(p_ws)
	store_in_excel(p_ws, t_ws)
	check_if_observed(p_ws, t_ws)
