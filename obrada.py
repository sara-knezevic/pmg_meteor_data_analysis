from openpyxl import *
import time, datetime
import os, shutil
import sameMeteor
import numpy as np

directory = "/home/sarasdfg/Petnica/Kodovi"
name = "Statistika2.xlsx"
date = "11-12"

sameMeteor.storingTime(directory, name, date)
sameMeteor.write(date)

timeFormat = "%Y-%m-%d %H:%M:%S"
wb = load_workbook("obradjivanje.xlsx")
sheet = wb.active

i = 0

for r in range(2, 2 + len(sameMeteor.times)):
	sheet.cell(row = r, column = 1).value = time.strftime(timeFormat, sameMeteor.times[i][1])

	mean = 0
	magnitudes = []

	for key, (met, mag) in sameMeteor.times[i][2].items():
		mean += mag

		magnitudes.append(mag)

	minMag = np.min(magnitudes)
	maxMag = np.max(magnitudes)


	sheet.cell(row = r, column = 2).value = round(mean / len(sameMeteor.times[i][2]), 3)
	sheet.cell(row = r, column = 3).value = minMag
	sheet.cell(row = r, column = 4).value = maxMag
	sheet.cell(row = r, column = 5).value = len(sameMeteor.times[i][2])

	i += 1

wb.save("obradjivanje.xlsx")
